import pandas as pd
import math
import locale

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from datetime import datetime
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment
from unidecode import unidecode


class Employee:
    cost: float
    name: str
    exams: str
    function: str
    data: str
    typeExam: str

    def __init__(self, name: str = '', cost: float = 0.0, function: str = '',
                 exams: str = '', data: str = '', typeExam: str = '',):
        self.name = name
        self.cost = cost
        self.function = function
        self.exams = exams
        self.data = data
        self.typeExam = typeExam


class Company:
    name: str
    exams: str
    missingExams: str
    employees: List[Employee]

    def __init__(self, name: str = '',
                 exams: str = '', employees: List[Employee] = [],
                 missingExams: str = ''):
        self.name = name
        self.missingExams = missingExams
        self.exams = exams
        self.employees = employees

    def __repr__(self):
        return f'Nome empresa: {self.name}, \n Lista de Exames: {self.exams},'\
            f'\n Exames Faltando: {self.missingExams})'

    def employeesList(self):
        employeeString = []
        for employee_person in self.employees:
            formatedCoast = math.ceil(employee_person.cost * 100)/100
            employeeString.append(f'Nome: {employee_person.name}, '
                                  f'Custo: {formatedCoast}')
        return employeeString


class EmployeesBilling:
    name: str
    function: str
    data: datetime
    exams: str
    typeExam: str

    def __init__(self, name: str = '', function: str = '',
                 exams: str = '',
                 data: datetime = datetime.now(), typeExam: str = ''):
        self.name = name
        self.function = function
        self.exams = exams
        self.data = data
        self.typeExam = typeExam


class CompanyBilling:
    name: str
    employeesBilling: List[EmployeesBilling]

    def __init__(self, name: str = '', employeesBilling=[]):
        self.name = name
        self.employeesBilling = employeesBilling


class Billing:
    dataExam: str
    nameCompany: str
    nameEmployees: str
    functionEmployees: str
    exams: str
    typeExam: str

    def __init__(self, dataExam: str = '', nameCompany: str = '',
                 nameEmployees: str = '', functionEmployees: str = '',
                 exams: str = '', typeExam: str = ''):
        self.nameCompany = nameCompany
        self.nameEmployees = nameEmployees
        self.functionEmployees = functionEmployees
        self.exams = exams
        self.dataExam = dataExam
        self.typeExam = typeExam


def getExams(company_name, list_of_company, newCompany):
    df = pd.read_excel(WORKBOOK_PATH_COMPANY, sheet_name=company_name)

    table = df.iloc[3:, 0:8]
    # tableOfCompany_ValueTeste = table.iloc[:, [0, 1]]
    # print(tableOfCompany_ValueTeste)
    tableOfCompany_Value = table.iloc[:, [0, 1]]
    # for teste in testeFull:
    for exam in tableOfCompany_Value.values.tolist():
        if (exam[0] is not None):
            list_of_company.append((exam[0], exam[1]))
            if (newCompany.exams != ''):
                newCompany.exams = newCompany.exams + ", " + str(exam[0])
            else:
                newCompany.exams = str(exam[0])


# ROOT_FOLDER = Path(__file__).parent
# WORKBOOK_PATH_COMPANY = ROOT_FOLDER / 'ValoresEmpresas.xlsx'
# WORKBOOK_PATH_FUNC = ROOT_FOLDER / 'examesRealizados.xlsx'
WORKBOOK_PATH_COMPANY = './ValoresEmpresas.xlsx'
WORKBOOK_PATH_FUNC = './examesRealizados.xlsx'

# Carregando um arquivo do excel
workbook_company: Workbook = load_workbook(WORKBOOK_PATH_COMPANY)
workbook_employees: Workbook = load_workbook(WORKBOOK_PATH_FUNC)

name_month = 'JUNHO'
sheet_name_employees = f'{name_month} 2023'

worksheet_employees: Worksheet = workbook_employees[sheet_name_employees]

billingList = []
for billing_row in worksheet_employees.iter_rows(min_row=2, values_only=True):
    if (billing_row[0] is not None):
        billingList.append(Billing(
            str(billing_row[0]), str(billing_row[1]),
            str(billing_row[2]), str(billing_row[3]),
            str(billing_row[4]), str(billing_row[5])))

companyList_Billing = []
for billing_row in billingList:
    # if (i > 50):
    #     continue
    noHasCompany = True
    for company in companyList_Billing:
        newCompanyBilling = CompanyBilling()
        if (billing_row.nameCompany == company.name):
            employeesBillingAux = EmployeesBilling(
                billing_row.nameEmployees,
                billing_row.functionEmployees,
                billing_row.exams,
                data=billing_row.dataExam,
                typeExam=billing_row.typeExam)
            company.employeesBilling.append(employeesBillingAux)
            noHasCompany = False
    if (noHasCompany):
        employeesBillingAux = EmployeesBilling(
            name=billing_row.nameEmployees,
            function=billing_row.functionEmployees,
            exams=billing_row.exams,
            data=billing_row.dataExam,
            typeExam=billing_row.typeExam)
        newCompanyBilling = CompanyBilling(
            name=billing_row.nameCompany,
            employeesBilling=[employeesBillingAux])
        companyList_Billing.append(newCompanyBilling)
    if (len(companyList_Billing) == 0):
        newCompanyBilling = CompanyBilling()

        employeesBillingAux = EmployeesBilling(
            name=billing_row.nameEmployees,
            function=billing_row.functionEmployees,
            exams=billing_row.exams,
            data=billing_row.dataExam,
            typeExam=billing_row.typeExam)
        newCompanyBilling = CompanyBilling(
            name=billing_row.nameCompany,
            employeesBilling=[employeesBillingAux])
        companyList_Billing.append(newCompanyBilling)

companyList_BillingTeste = []
for i in range(10):
    companyList_BillingTeste.append(companyList_Billing[i])

names_of_workbook = workbook_company.sheetnames

namesOfCompanys = []
for names in names_of_workbook:
    ws: Worksheet = workbook_company[f'{names}']
    namesOfCompanys.append((ws['B1'].value, names))
# companyList_BillingTeste = [companyList_Billing[9]]
companys_not_found = 'Empresas não encontradas: '
companyList = []
for company_Billing in companyList_BillingTeste:

    newCompany = Company(company_Billing.name)
    newCompany.employees = []

    company_name_billing = unidecode(company_Billing.name.lower())
    hasName = False
    company_name = ''
    for names in namesOfCompanys:
        # print('Empresa que tenho o sheet', unidecode(names[0]),
        #       '  | Empresa que ta na tabela de exames:', company_name_billing)
        if (unidecode(names[0].lower()) in company_name_billing):
            # Selecionou a planilha da empresa'
            company_name = names[1]
            hasName = True
            break
    if (not hasName):
        companys_not_found = companys_not_found + ', \n' + company_Billing.name
    else:
        list_of_company = []
        getExams(company_name, list_of_company, newCompany)
        names_exams = []
        index = 0
        employees_cost = []

        for employees in company_Billing.employeesBilling:
            names_exams.append(
                (employees.name, employees.exams, employees.function,
                 employees.exams, employees.data, employees.typeExam))

        for nameExam in names_exams:
            data_str_fmt_for_date = '%Y-%m-%d %H:%M:%S'
            data_str_fmt = '%d/%m/%Y '
            date_Typed = datetime.strptime(nameExam[4], data_str_fmt_for_date)
            dateFormatted = date_Typed.strftime(data_str_fmt)
            employee_company = Employee()
            employee_company.name = nameExam[0]
            employee_company.function = nameExam[2]
            employee_company.exams = nameExam[3]
            employee_company.typeExam = nameExam[5]
            employee_company.data = dateFormatted
            exams_exect = nameExam[1].split('/')

            hasExam = False
            for exam in exams_exect:
                exam = exam.strip()
                if (exam == 'CLINICO'):
                    exam = 'Clínico'
                if (exam == 'AUDIO'):
                    exam = 'Audiometria'
                elif (exam == "HEMO"):
                    exam = 'Hemograma c/ Plaquetas'
                elif (exam == 'AC HIPURICO'):
                    exam = 'Ácido Hipúrico'
                elif (exam == 'AC METIL'):
                    exam = 'Ácido Metil Hipúrico'
                elif (exam == 'AC METIL HIPURICO'):
                    exam = 'Ácido Metil Hipúrico'
                elif (exam == 'AC MANDELICO'):
                    exam = 'Ácido Mandélico'
                elif (exam == 'RX DE TORAX'):
                    exam = 'Raio-x de Tórax'
                elif (exam == 'RX DE TORAX'):
                    exam = 'Raio-x de Tórax'
                elif (exam == 'AV PSICOSSOCIAL'):
                    exam = 'Avaliação Psocossocial'
                hasExam = False
                for company in list_of_company:
                    if (unidecode(exam.lower())
                            in unidecode(company[0].lower())
                            and not hasExam):
                        if (('externo' not in exam.lower() and
                                'externo' not in company[0].lower())):
                            employee_company.cost += company[1]
                            hasExam = True
                if (not hasExam):
                    if (newCompany.missingExams == ''):
                        auxStr = "ah empresa não tem o(s) exame(s): "
                        newCompany.missingExams = auxStr + exam

                    elif (not (exam in newCompany.missingExams)):
                        newCompany.missingExams = newCompany.missingExams + ""\
                            ", " + exam
            newCompany.employees.append(employee_company)
        companyList.append(newCompany)

for company in companyList:
    print("========================================================")
    print("COMPANY LIST: \n",  company)
    print("COMPANY LIST OF EMPLOYEES: \n", company.employeesList())
    print("========================================================")
print(companys_not_found)

columns = ['Mês', 'Nome', 'Função', 'Exames', 'Tipo de Exame', 'Valor']
wb: Workbook = Workbook()
wb.remove(wb['Sheet'])
locale.setlocale(locale.LC_ALL, 'pt_BR')
mnum = datetime.strptime(name_month, '%B').month

# Estilo da tabela
style_header = NamedStyle(name="style_header")
style_header.font = Font(size=12, color='FFFFFF', name='Arial', )
style_header.font.bold = True
style_header.alignment = Alignment(horizontal="center", vertical="center")
style_header.fill = PatternFill("solid", fgColor="76933c")
style_content = NamedStyle(name="style_content")
style_content.font = Font(size=10, name='Arial')
style_content.fill = PatternFill("solid", fgColor="d8e4bc")

wb.add_named_style(style_header)
wb.add_named_style(style_content)

for company in companyList:
    dataCompany = []
    for employeesItem in company.employees:
        formatedCoast = math.ceil(employeesItem.cost * 100)/100
        aux = [employeesItem.data,
               employeesItem.name,
               employeesItem.function,
               employeesItem.exams,
               employeesItem.typeExam,
               formatedCoast]
        dataCompany.append(aux)
    df = pd.DataFrame(dataCompany, columns=columns)

    ws: Worksheet = wb.create_sheet(company.name[:15])

    # Montando a tabela

    # Header
    ws.append([f'Mês {mnum:02}/23',
               company.name, '', '', '', 'Valor'])

    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']

    ws.merge_cells('B1:E1')

    for item in dataCompany:
        ws.append(item)
    # Header Styles
    for i in columnsTable:
        ws[f'{i}1'].style = style_header

    # Conteúdo da tabela
    for i in range(0, len(dataCompany)):
        for y in columnsTable:
            ws[f'{y}{i+2}'].style = style_content
            if (y == 'A'):
                ws[f'{y}{i+2}'].alignment = Alignment(
                    horizontal="center", vertical="center")
            if (y == 'F'):
                number_forma = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                ws[f'{y}{i+2}'].number_format = number_forma
    # Redimensionar colunas da tabela
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4

wb.save('Faturamento_Empresas.xlsx')
# ----------------------------------------------------------------

# df.to_excel("Faturamento_teste.xlsx", sheet_name="Plan1",
#             index=False, engine='openpyxl')
