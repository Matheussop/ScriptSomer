import pandas as pd
import math
import locale
import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from datetime import datetime
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment, Border
from openpyxl.styles import Side
from unidecode import unidecode

dictionary_exams = [
    ("CLINICO", "Clínico"),   ("AUDIO", "Audiometria"),
    ("HEMO", "Hemograma c/ Plaquetas"), ("AC METIL", "Ácido Metil Hipúrico"),
    ("AC HIPURICO", "Ácido Hipúrico"), ("AC MANDELICO", "Ácido Mandélico"),
    ("AC METIL HIPURICO", "Ácido Metil Hipúrico"),
    ("RX DE TORAX", "Raio-x de Tórax"),
    ("AV PSICOSSOCIAL", "Avaliação Psicossocial"),
    ("CLINICO EXTERNO", "Exame Clínico (in loco)")
]

dictionary_company_names = [
    ("Minas Brasil", "BRASIL COMERCIAL"),
    ("ACTur", "AC TRANSPORTES"),
]


class Employee:
    cost: float
    name: str
    exams: str
    function: str
    data: str
    examsCost: List
    typeExam: str

    def __init__(self, examsCost: List = [], name: str = '', cost: float = 0.0,
                 function: str = '', exams: str = '', data: str = '',
                 typeExam: str = '',):
        self.name = name
        self.cost = cost
        self.function = function
        self.exams = exams
        self.data = data
        self.typeExam = typeExam
        self.examsCost = examsCost


class Company:
    name: str
    exams: str
    missingExams: str
    employees: List[Employee]

    def __init__(self, name: str = '',
                 exams: str = '', employees: List[Employee] = [],
                 missingExams: str = ''):
        self.name = name
        self.missingExams = missingExams
        self.exams = exams
        self.employees = employees

    def __repr__(self):
        return f'Nome empresa: {self.name}, \n Lista de Exames: {self.exams},'\
            f'\n Exames Faltando: {self.missingExams})'

    def employeesList(self):
        employeeString = []
        for employee_person in self.employees:
            formatedCoast = math.ceil(employee_person.cost * 100)/100
            employeeString.append(f'Nome: {employee_person.name}, '
                                  f'Custo: {formatedCoast}')
        return employeeString


class EmployeesBilling:
    name: str
    function: str
    data: datetime
    exams: str
    typeExam: str

    def __init__(self, name: str = '', function: str = '',
                 exams: str = '',
                 data: datetime = datetime.now(), typeExam: str = ''):
        self.name = name
        self.function = function
        self.exams = exams
        self.data = data
        self.typeExam = typeExam


class CompanyBilling:
    name: str
    employeesBilling: List[EmployeesBilling]

    def __init__(self, name: str = '', employeesBilling=[]):
        self.name = name
        self.employeesBilling = employeesBilling


class Billing:
    dataExam: str
    nameCompany: str
    nameEmployees: str
    functionEmployees: str
    exams: str
    typeExam: str

    def __init__(self, dataExam: str = '', nameCompany: str = '',
                 nameEmployees: str = '', functionEmployees: str = '',
                 exams: str = '', typeExam: str = ''):
        self.nameCompany = nameCompany
        self.nameEmployees = nameEmployees
        self.functionEmployees = functionEmployees
        self.exams = exams
        self.dataExam = dataExam
        self.typeExam = typeExam


# ROOT_FOLDER = Path(__file__).parent
# WORKBOOK_PATH_COMPANY = ROOT_FOLDER / 'ValoresEmpresas.xlsx'
# WORKBOOK_PATH_FUNC = ROOT_FOLDER / 'examesRealizados.xlsx'
WORKBOOK_PATH_COMPANY = './ValoresEmpresas.xlsx'
WORKBOOK_PATH_FUNC = './examesRealizados.xlsx'
hasDetailed = False


def getExams(company_name, list_of_company, newCompany):
    df = pd.read_excel(WORKBOOK_PATH_COMPANY, sheet_name=company_name)

    table = df.iloc[3:, 0:8]
    # tableOfCompany_ValueTeste = table.iloc[:, [0, 1]]
    # print(tableOfCompany_ValueTeste)
    tableOfCompany_Value = table.iloc[:, [0, 1]]
    # for teste in testeFull:
    for exam in tableOfCompany_Value.values.tolist():
        if (exam[0] is not None):
            list_of_company.append((exam[0], exam[1]))
            if (newCompany.exams != ''):
                newCompany.exams = newCompany.exams + ", " + str(exam[0])
            else:
                newCompany.exams = str(exam[0])


def setDateXmlsToString(dataInXmls) -> str:
    # Bloco para fixar a data como string
    data_str_fmt_for_date = '%Y-%m-%d %H:%M:%S'
    data_str_fmt = '%d/%m/%Y '
    date_Typed = datetime.strptime(dataInXmls, data_str_fmt_for_date)
    return date_Typed.strftime(data_str_fmt)


def styleXmls(wb):
    # Estilo da tabela
    bd = Side(style='thin', color="000000")
    style_header = NamedStyle(name="style_header")
    style_header.font = Font(size=12, color='FFFFFF', name='Arial', )
    style_header.font.bold = True
    style_header.alignment = Alignment(horizontal="center", vertical="center")
    style_header.fill = PatternFill("solid", fgColor="76933c")
    style_header.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    style_content = NamedStyle(name="style_content")
    style_content.font = Font(size=10, name='Arial')
    style_content.fill = PatternFill("solid", fgColor="d8e4bc")
    style_content.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    style_Cost = NamedStyle(name="style_Cost")
    style_Cost.font = Font(size=12, name='Arial', )
    style_Cost.font.bold = True
    style_Cost.alignment = Alignment(horizontal="center", vertical="center")
    style_Cost.fill = PatternFill("solid", fgColor="d8e4bc")
    style_Cost.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    wb.add_named_style(style_header)
    wb.add_named_style(style_content)
    wb.add_named_style(style_Cost)


def setContentDataFrame(arrayUnformattedData: List) -> List:
    data = []
    for employeesItem in arrayUnformattedData:
        formattedCoast = math.ceil(employeesItem.cost * 100)/100
        aux = [employeesItem.data,
               employeesItem.name,
               employeesItem.function,
               employeesItem.exams,
               employeesItem.typeExam,
               formattedCoast]
        data.append(aux)
    return data


def setContentDetailDataFrame(arrayUnformattedData: List) -> List:
    data = []
    for employeesItem in arrayUnformattedData:
        listOfExams = []
        listOfCost = []
        for examCost in employeesItem.examsCost:
            formattedCoastDetail = math.ceil(examCost[1] * 100)/100
            listOfExams.append(examCost[0])  # name exam
            listOfCost.append(formattedCoastDetail)  # value exam

        aux = [employeesItem.data,
               employeesItem.name,
               employeesItem.function,
               listOfExams,
               employeesItem.typeExam,
               listOfCost]
        data.append(aux)
    return data


def mountStyleHeaderTable(monthNumber: int, ws: Worksheet, companyName: str):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']
    # Header
    ws.append([f'Mês {monthNumber:02}/23',
               companyName, '', '', '', 'Valor'])
    # Aplicando Header Styles
    for i in columnsTable:
        ws[f'{i}1'].style = 'style_header'

    ws.merge_cells('B1:E1')


def mountStyleContentTable(ws, dataCompany):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']
    count_row = ws.max_row

    # Conteúdo da tabela
    for i in range(0, count_row-1):
        for y in columnsTable:
            ws[f'{y}{i+2}'].style = 'style_content'
            # if (y == 'A'):
            #     ws[f'{y}{i+2}'].alignment = Alignment(
            #         horizontal="center", vertical="center")
            ws[f'{y}{i+2}'].alignment = Alignment(
                horizontal="center", vertical="center")
            if (y == 'F'):
                number_forma = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                ws[f'{y}{i+2}'].number_format = number_forma
                ws[f'{y}{i+2}'].alignment = Alignment(
                    horizontal="right", vertical="center")
            if (y == 'D'):
                ws[f'{y}{i+2}'].alignment = Alignment(
                    horizontal="left", vertical="center")


def resizeTable(ws):
    # Redimensionar colunas da tabela
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4


def calculateCost(company) -> int:
    cost = 0
    for employee in company.employees:
        cost += employee.cost
    return cost


def setTotalCostStyle(ws):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']
    row_count = ws.max_row
    number_forma = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    for letter in columnsTable:
        ws[f'{letter}{row_count}'].style = 'style_Cost'

    ws[f'F{row_count}'].number_format = number_forma


def setTotalCost(ws):
    row_count = ws.max_row
    ws.append(['TOTAL', '', '', '', '', f"=SUM(F2: F{row_count})"])
    ws.merge_cells(f'A{row_count+1}:E{row_count+1}')
    setTotalCostStyle(ws)


def addDetailDataFrame(ws, item):
    row_count = ws.max_row
    lenOfExams = len(item[3])
    exams = item[3]
    examsCost = item[5]
    firstRow = [item[0], item[1], item[2], exams[0], item[4], examsCost[0]]
    ws.append(firstRow)
    for i in range(1, lenOfExams, ):
        ws.append(['', '', '',  exams[i], '', examsCost[i]])

    columnsAppend = ['A', 'B', 'C', 'E']
    # merge_cells A, B, C, E
    for letter in columnsAppend:
        ws.merge_cells(
            f'{letter}{row_count+1}:{letter}{row_count+lenOfExams}')


def main():
    # Carregando um arquivo do excel
    workbook_company: Workbook = load_workbook(WORKBOOK_PATH_COMPANY)
    workbook_employees: Workbook = load_workbook(WORKBOOK_PATH_FUNC)

    name_month = 'JUNHO'
    sheet_name_employees = f'{name_month} 2023'

    worksheet_employees: Worksheet = workbook_employees[sheet_name_employees]

    billingList = []
    for billing_row in worksheet_employees.iter_rows(min_row=2,
                                                     values_only=True):
        if (billing_row[0] is not None):
            billingList.append(Billing(
                str(billing_row[0]), str(billing_row[1]),
                str(billing_row[2]), str(billing_row[3]),
                str(billing_row[4]), str(billing_row[5])))

    companyList_Billing = []
    for billing_row in billingList:
        # if (i > 50):
        #     continue
        noHasCompany = True
        for company in companyList_Billing:
            newCompanyBilling = CompanyBilling()
            if (billing_row.nameCompany == company.name):
                employeesBillingAux = EmployeesBilling(
                    billing_row.nameEmployees,
                    billing_row.functionEmployees,
                    billing_row.exams,
                    data=billing_row.dataExam,
                    typeExam=billing_row.typeExam)
                company.employeesBilling.append(employeesBillingAux)
                noHasCompany = False
        if (noHasCompany):
            employeesBillingAux = EmployeesBilling(
                name=billing_row.nameEmployees,
                function=billing_row.functionEmployees,
                exams=billing_row.exams,
                data=billing_row.dataExam,
                typeExam=billing_row.typeExam)
            newCompanyBilling = CompanyBilling(
                name=billing_row.nameCompany,
                employeesBilling=[employeesBillingAux])
            companyList_Billing.append(newCompanyBilling)
        if (len(companyList_Billing) == 0):
            newCompanyBilling = CompanyBilling()
            employeesBillingAux = EmployeesBilling(
                name=billing_row.nameEmployees,
                function=billing_row.functionEmployees,
                exams=billing_row.exams,
                data=billing_row.dataExam,
                typeExam=billing_row.typeExam)
            newCompanyBilling = CompanyBilling(
                name=billing_row.nameCompany,
                employeesBilling=[employeesBillingAux])
            companyList_Billing.append(newCompanyBilling)

    # Bloco para testar numero x de empresas
    companyList_BillingTeste = []
    for i in range(20):
        companyList_BillingTeste.append(companyList_Billing[i])

    # companyList_BillingTeste = [companyList_Billing[3]]
    # ----------------------------------------------------

    # Percorrer o sheet e pegar o nome da empresa dentro da tabela
    names_of_workbook = workbook_company.sheetnames
    namesOfCompanys = []
    for names in names_of_workbook:
        ws: Worksheet = workbook_company[f'{names}']
        namesOfCompanys.append((ws['B1'].value, names))

    for nameCompany in dictionary_company_names:
        namesOfCompanys.append((nameCompany[1], nameCompany[0]))
    # ----------------------------------------------------
    companys_not_found = 'Empresas não encontradas: '
    companyList = []
    for company_Billing in companyList_BillingTeste:

        newCompany = Company(company_Billing.name)
        newCompany.employees = []

        company_name_billing = unidecode(company_Billing.name.lower())
        hasName = False
        company_name = ''
        for names in namesOfCompanys:
            namesAux = unidecode(names[0].lower())
            namesAux2 = unidecode(names[1].lower())
            company_name_billing = company_name_billing.replace('ltda', '')
            company_name_billing = company_name_billing.replace('eireli', '')
            namesAux = namesAux.replace('ltda', '')
            namesAux = namesAux.replace('eireli', '')
            # print('Empresa que tenho o sheet', namesAux,
            #       '  | Empresa que ta na tabela de exames:', company_name_billing)

            if ((namesAux in company_name_billing)
                    or (namesAux2 in company_name_billing)):
                company_name = names[1]
                hasName = True
                break
        if (not hasName):
            companys_not_found = companys_not_found + ', \n' + company_Billing.name
        else:
            list_of_company = []
            getExams(company_name, list_of_company, newCompany)
            namesAndExams = []

            for employees in company_Billing.employeesBilling:
                namesAndExams.append(
                    (employees.name, employees.exams, employees.function,
                     employees.exams, employees.data, employees.typeExam))

            for nameAndExam in namesAndExams:
                dateFormatted = setDateXmlsToString(nameAndExam[4])
                employee_company = Employee()
                employee_company.name = nameAndExam[0]
                employee_company.function = nameAndExam[2]
                employee_company.exams = nameAndExam[3]
                employee_company.typeExam = nameAndExam[5]
                employee_company.data = dateFormatted
                employee_company.examsCost = []
                exams_exect = nameAndExam[1].split('/')

                hasExam = False
                for exam in exams_exect:
                    exam = exam.strip()
                    exam_compar = unidecode(exam.lower())
                    for exam_significant in dictionary_exams:
                        examDictionary = unidecode(exam_significant[0].lower())
                        if (exam_compar == examDictionary):
                            exam = exam_significant[1]

                    hasExam = False
                    for company in list_of_company:
                        if (unidecode(exam.lower())
                                in unidecode(company[0].lower())
                                and not hasExam):
                            if (('externo' not in exam.lower() and
                                    'externo' not in company[0].lower())):
                                employee_company.cost += company[1]
                                if (hasDetailed):
                                    employee_company.examsCost.append(
                                        (exam, company[1]))
                                hasExam = True
                    if (not hasExam):
                        if (newCompany.missingExams == ''):
                            auxStr = "ah empresa não tem o(s) exame(s): "
                            newCompany.missingExams = auxStr + exam

                        elif (not (exam in newCompany.missingExams)):
                            newCompany.missingExams = newCompany.missingExams + ""\
                                ", " + exam
                newCompany.employees.append(employee_company)
            companyList.append(newCompany)

    # for company in companyList:
    #     print("========================================================")
    #     print("COMPANY LIST: \n",  company)
    #     # print("COMPANY LIST OF EMPLOYEES: \n", company.employeesList())
    #     print("========================================================")
    # print(companys_not_found)

    if len(companyList) > 0:

        locale.setlocale(locale.LC_ALL, 'pt_BR')
        monthNumber = datetime.strptime(name_month, '%B').month
        try:
            os.mkdir('Faturamentos')
        except FileExistsError:
            ...
        except Exception as e:
            print('Error ao criar a pasta', e)

        for company in companyList:
            wb: Workbook = Workbook()
            styleXmls(wb)
            wb.remove(wb['Sheet'])
            dataCompany = []

            if (hasDetailed):
                dataCompany = setContentDetailDataFrame(company.employees)
            else:
                dataCompany = setContentDataFrame(company.employees)

            ws: Worksheet = wb.create_sheet(company.name[:15])

            mountStyleHeaderTable(monthNumber, ws, company.name)

            # cost = 0
            # cost = calculateCost(company)
            # print(cost)
            for item in dataCompany:
                if (hasDetailed):
                    addDetailDataFrame(ws, item)
                else:
                    ws.append(item)

            mountStyleContentTable(ws, dataCompany)
            setTotalCost(ws)
            resizeTable(ws)

            nameFile = f'{company.name} - {monthNumber} {name_month}'
            wb.save(f'Faturamentos/Faturamento - {nameFile} 2023.xlsx')

# ----------------------------------------------------------------


if __name__ == '__main__':
    input = input("Deseja que a planilha seja detalhada ? (S/N)")
    if (input.lower() == "s"):
        hasDetailed = True
    main()
