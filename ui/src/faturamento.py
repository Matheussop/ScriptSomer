import pandas as pd
import math
import json
import locale
import sys
import asyncio
import os
import logging

from PySide6.QtCore import QObject, Signal
from argparse import ArgumentParser
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Tuple
from datetime import datetime
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment, Border
from openpyxl.styles import Side
from unidecode import unidecode

dictionary_company_names = [
    ("Minas Brasil", "BRASIL COMERCIAL"),
    ("ACTur", "AC TRANSPORTES"),
]
is_detail = False


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(
        os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


WORKBOOK_PATH_COMPANY = resource_path('./files/ValoresEmpresas.xlsx')
WORKBOOK_PATH_FUNC = resource_path('./files/examesRealizados.xlsx')


class ErrorBilling(Exception):
    ...


class Employee:
    cost: float
    name: str
    exams: str
    function: str
    date: str
    exams_cost: List
    exam_type: str

    def __init__(self, exams_cost: List = [], name: str = '',
                 cost: float = 0.0, function: str = '',
                 exams: str = '', date: str = '', exam_type: str = '',):
        self.name = name
        self.cost = cost
        self.function = function
        self.exams = exams
        self.date = date
        self.exam_type = exam_type
        self.exams_cost = exams_cost


class Company:
    name: str
    exams: str
    missingExams: str
    employees: List[Employee]

    def __init__(self, name: str = '',
                 exams: str = '', employees: List[Employee] = [],
                 missingExams: str = ''):
        self.name = name
        self.missingExams = missingExams
        self.exams = exams
        self.employees = employees

    def __repr__(self):
        return f'Nome empresa: {self.name}, \n Lista de Exames: {self.exams},'\
            f'\n Exames Faltando: {self.missingExams})'

    def employeesList(self):
        employeeString = []
        for employee_person in self.employees:
            formattedCoast = math.ceil(employee_person.cost * 100)/100
            employeeString.append(f'Nome: {employee_person.name}, '
                                  f'Custo: {formattedCoast}')
        return employeeString


class EmployeesBilling:
    name: str
    function: str
    data: datetime
    exams: str
    exam_type: str

    def __init__(self, name: str = '', function: str = '',
                 exams: str = '',
                 data: datetime = datetime.now(), exam_type: str = ''):
        self.name = name
        self.function = function
        self.exams = exams
        self.data = data
        self.exam_type = exam_type


class CompanyBilling:
    name: str
    employeesBilling: List[EmployeesBilling]

    def __init__(self, name: str = '', employeesBilling=[]):
        self.name = name
        self.employeesBilling = employeesBilling


class Billing:
    dataExam: str
    company_name: str
    employee_name: str
    employee_function: str
    exams: str
    exam_type: str

    def __init__(self, dataExam: str = '', company_name: str = '',
                 employee_name: str = '', employee_function: str = '',
                 exams: str = '', exam_type: str = ''):
        self.company_name = company_name
        self.employee_name = employee_name
        self.employee_function = employee_function
        self.exams = exams
        self.dataExam = dataExam
        self.exam_type = exam_type


def getExams(company_name, list_of_company_exams, newCompany):
    """
    Carrega os exames para uma empresa específica a partir do arquivo .xlsx.

    Args:
        company_name (str): Nome da empresa.
        list_of_company_exams (list): Lista para armazenar os exames.
        newCompany (Company): Objeto Company para armazenar os dados da
        empresa.

    Raises:
        ErrorBilling: Se ocorrer um erro ao acessar o arquivo.
    """
    ROW_START = 3
    COLUMN_INDEX = [0, 1]
    try:

        df = pd.read_excel(WORKBOOK_PATH_COMPANY, sheet_name=company_name)
        # Get row 3 to infinity and Column A and B
        tableOfCompany_Value = df.iloc[ROW_START:, COLUMN_INDEX]
        exams = []
        for exam in tableOfCompany_Value.values.tolist():
            if pd.isna(exam[0]):
                break
            if (exam[0] is not None):
                exams.append((exam[0], exam[1]))
                if (newCompany.exams != ''):
                    newCompany.exams = newCompany.exams + ", " + str(exam[0])
                else:
                    newCompany.exams = str(exam[0])
        list_of_company_exams.extend(exams)
    except FileNotFoundError:
        exception_ = ErrorBilling(
            f'Error ao tentar acessar o arquivo: {WORKBOOK_PATH_COMPANY}')
        raise exception_


def setDateXlsxToString(dataInXlsx) -> str:
    # Bloco para fixar a data como string
    data_str_fmt_for_date = '%Y-%m-%d %H:%M:%S'
    data_str_fmt = '%d/%m/%Y '
    date_Typed = datetime.strptime(dataInXlsx, data_str_fmt_for_date)
    return date_Typed.strftime(data_str_fmt)


def styleXlsx(wb):
    # Estilo da tabela
    bd = Side(style='thin', color="000000")
    style_header = NamedStyle(name="style_header")
    style_header.font = Font(size=12, color='FFFFFF', name='Arial', )
    style_header.font.bold = True
    style_header.alignment = Alignment(horizontal="center", vertical="center")
    style_header.fill = PatternFill("solid", fgColor="76933c")
    style_header.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    style_content = NamedStyle(name="style_content")
    style_content.font = Font(size=10, name='Arial')
    style_content.fill = PatternFill("solid", fgColor="d8e4bc")
    style_content.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    style_Cost = NamedStyle(name="style_Cost")
    style_Cost.font = Font(size=12, name='Arial', )
    style_Cost.font.bold = True
    style_Cost.alignment = Alignment(horizontal="center", vertical="center")
    style_Cost.fill = PatternFill("solid", fgColor="d8e4bc")
    style_Cost.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    wb.add_named_style(style_header)
    wb.add_named_style(style_content)
    wb.add_named_style(style_Cost)


def setContentDataFrame(array_unformatted_data: List[Employee]) -> List:
    data = []
    for employees_item in array_unformatted_data:
        aux = [employees_item.date,
               employees_item.name,
               employees_item.function,
               employees_item.exams,
               employees_item.exam_type,
               employees_item.cost]
        data.append(aux)
    return data


def setContentDetailDataFrame(array_unformatted_data: List[Employee]) -> List:
    data = []
    for employees_item in array_unformatted_data:
        list_of_exams = []
        list_of_cost = []
        for exam_cost in employees_item.exams_cost:
            list_of_exams.append(exam_cost[0])  # Nome do exame
            list_of_cost.append(exam_cost[1])   # Valor do exame

        aux = [employees_item.date,
               employees_item.name,
               employees_item.function,
               list_of_exams,
               employees_item.exam_type,
               list_of_cost]
        data.append(aux)
    return data


def mountStyleHeaderTable(yearText: str, monthNumber: int, ws: Worksheet,
                          companyName: str):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']

    # Cabeçalho da planilha
    ws.append([f'Mês {monthNumber:02}/{int(yearText[2:])}',
               companyName, '', '', '', 'Valor'])

    # Aplicando Header Styles
    for i in columnsTable:
        ws[f'{i}1'].style = 'style_header'

    # Unindo a células onde ficara o nome da empresa
    ws.merge_cells('B1:E1')


def mountStyleContentTable(ws):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']
    count_row = ws.max_row

    # Definindo o estilo e formatação do conteúdo da tabela
    for i in range(0, count_row-1):
        for y in columnsTable:
            ws[f'{y}{i+2}'].style = 'style_content'

            ws[f'{y}{i+2}'].alignment = Alignment(
                horizontal="center", vertical="center")
            if (y == 'F'):
                # Formula para forçar a ser um campo de dinheiro "BR"
                number_forma = '"$"#,##0.00_);[Red]("$"#,##0.00)'
                ws[f'{y}{i+2}'].number_format = number_forma
                ws[f'{y}{i+2}'].alignment = Alignment(
                    horizontal="right", vertical="center")
            if (y == 'D'):
                ws[f'{y}{i+2}'].alignment = Alignment(
                    horizontal="left", vertical="center")


def resizeTable(ws):
    # Redimensionar colunas da tabela
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 4


def calculateCost(company) -> int:
    cost = 0
    for employee in company.employees:
        cost += employee.cost
    return cost


def setTotalCostStyle(ws):
    columnsTable = ['A', 'B', 'C', 'D', 'E', 'F']
    row_count = ws.max_row
    # Formula para forçar a ser um campo de dinheiro "BR"
    number_forma = '"$"#,##0.00_);[Red]("$"#,##0.00)'
    for letter in columnsTable:
        ws[f'{letter}{row_count}'].style = 'style_Cost'

    ws[f'F{row_count}'].number_format = number_forma


def setTotalCostRow(ws):
    row_count = ws.max_row
    # Espaços em branco pois será onde a célula sofrera um merge
    ws.append(['TOTAL', '', '', '', '', f"=SUM(F2: F{row_count})"])
    ws.merge_cells(f'A{row_count+1}:E{row_count+1}')
    setTotalCostStyle(ws)


def addDetailDataFrame(ws, item):
    row_count = ws.max_row
    # Tamanho da lista de exames feitos pelo funcionário
    len_of_exams = len(item[3])
    exams = item[3]
    exams_cost = item[5]
    # Caso não tenha nenhum exame limpa o campo.
    if (exams == []):
        exams.append("-")
        exams_cost.append("-")

    first_row = [item[0], item[1], item[2], exams[0], item[4], exams_cost[0]]
    ws.append(first_row)
    for i in range(1, len_of_exams):
        # Espaços em branco pois será onde a célula sofrera um merge
        ws.append(['', '', '',  exams[i], '', exams_cost[i]])

    columns_merge = ['A', 'B', 'C', 'E']
    # merge_cells A, B, C, E
    if (len(exams) >= 2):
        for letter in columns_merge:
            ws.merge_cells(
                f'{letter}{row_count+1}:{letter}{row_count+len_of_exams}')


def showMissingExams(ws, missingExams):
    text = 'Esta empresa não tem o(s) seguintes exame(s) cadastrados: '
    ws.append([f'{text}', '', f'{missingExams}',
              '', '', ''])
    ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
    ws.merge_cells(f'C{ws.max_row}:F{ws.max_row}')


def showCompanyInTerminal(companyList: List[Company]):
    for company in companyList:
        print("========================================================")
        print("COMPANY LIST: \n",  company)
        print("COMPANY LIST OF EMPLOYEES: \n", company.employeesList())
        print("========================================================")


def showInTerminalCompanyNameNumber(companyName, number):
    print('Numero: ', number,
          ' Nome empresa', companyName)


class BillingDataProcessor(QObject):
    """
    Classe responsável por processar o faturamento dos exames médicos
    realizados por empresas.

    Atributos:
        yearText (str): Ano do faturamento.
        monthText (str): Mês do faturamento.
        folder (str): Pasta onde os arquivos gerados serão salvos.
        is_detail (bool): Indica se os arquivos devem ser detalhados.
        dictionary_exams (list): Dicionário com os exames e seus valores.
        maxEmployees (int): Número máximo de funcionários.
        worksheet_employees (Worksheet): Planilha de exames realizados pelos
        funcionarios.
        workbook_employees (Workbook): Workbook de exames realizados.
        names_of_companies (list): Lista de nomes de empresas.
        companies_not_found (list): Lista de empresas não encontradas.
        companyList (list): Lista de empresas.
        companyList_Billing (list): Lista de empresas para faturamento.
        started (Signal): Sinal emitido quando o processo é iniciado.
        progressed (Signal): Sinal emitido quando o progresso é atualizado.
        finished (Signal): Sinal emitido quando o processo é finalizado.
        range_progress (Signal): Sinal emitido para definir o intervalo de
        progresso.
        companies_not_found_signal (Signal): Sinal emitido quando empresas não
        são encontradas.
        progress (int): Progresso atual do processo.
    """

    yearText = ''
    monthText = ''
    folder = ''
    is_detail = False
    dictionary_exams = []
    maxEmployees = 0
    worksheet_employees: Worksheet
    workbook_employees: Workbook
    names_of_companies = []
    companies_not_found = []
    companyList = []
    companyList_Billing = []
    started = Signal(str)
    progressed = Signal(int)
    finished = Signal(list)
    range_progress = Signal(int)
    companies_not_found_signal = Signal(list)
    progress = 0

    def __init__(self, yearText='', monthText='',
                 folder='', is_detail=False,
                 dictionary_exams=[],
                 names_of_companies=[],
                 parent=None) -> None:
        super().__init__(parent)
        self.yearText = yearText
        self.monthText = monthText
        self.folder = folder
        self.is_detail = is_detail
        self.dictionary_exams = dictionary_exams
        self.names_of_companies = names_of_companies
        self.maxEmployees = 0
        self.progress = 0

    def reset(self):
        """
        Reseta o estado do processo de faturamento.
        """
        self.maxEmployees = 0
        self.worksheet_employees = None  # type: ignore
        self.workbook_employees = None  # type: ignore
        self.names_of_companies = []
        self.companies_not_found = []
        self.companyList = []
        self.companyList_Billing = []
        self.progress = 0

    def setParamsBilling(self, yearText, monthText, folder,
                         is_detail) -> None:
        self.yearText = yearText
        self.monthText = monthText
        self.folder = folder
        self.is_detail = is_detail

    def load_exam_data(self):
        """
        Carrega os valores dos exames a partir do arquivo .xlsx.
        """
        try:
            # Carregando o arquivo de empresas do excel
            self.workbook_employees: Workbook = load_workbook(
                WORKBOOK_PATH_FUNC)
            logging.info('Arquivo de exames carregado com sucesso.')
        except Exception as e:
            exception_ = ErrorBilling(
                f'Error ao tentar acessar o arquivo: {WORKBOOK_PATH_FUNC} ')
            logging.error(f'Erro ao carregar o arquivo de exames: {e}')
            raise exception_

        if self.workbook_employees:
            name_month = self.monthText.upper()
            year = str(self.yearText)
            sheet_name_employees = f'{name_month} {year}'

            # Carregando o arquivo de exames realizados do excel
            try:
                self.worksheet_employees: Worksheet = \
                    self.workbook_employees[sheet_name_employees]
            except Exception:
                error = f'Error ao tentar encontrar '\
                    'uma página da planilha chamada ' \
                    f'{sheet_name_employees}'
                exception_ = ErrorBilling(error)

                raise exception_

    def getCompaniesNotFound(self): return self.companies_not_found

    def extract_billing_list(self):
        """
        Extrai a lista de faturamento do arquivo de entrada.

        Returns:
            List[Billing]: Lista de objetos Billing extraídos do arquivo.
        """
        billing_list = []
        for billing_row in self.worksheet_employees.\
                iter_rows(min_row=2, values_only=True):
            if billing_row[0] is not None:
                billing_list.append(Billing(
                    str(billing_row[0]),  # Data exame
                    str(billing_row[1]),  # Name company
                    str(billing_row[2]),  # Name employee
                    str(billing_row[3]),  # Function Employee
                    str(billing_row[4]),  # Exams
                    str(billing_row[5])   # Type Exam
                ))
        return billing_list

    def get_name_company_in_sheet(self) -> List[tuple]:
        """
        Retorna uma lista de tuplas contendo os nomes das empresas
        no workbook e seus respectivos nomes de sheet.

        Returns:
            List[Tuple[str, str]]: Lista de tuplas (nome_empresa, nome_sheet)
        """
        try:
            workbook_company: Workbook = load_workbook(WORKBOOK_PATH_COMPANY)
        except Exception:
            exception_ = ErrorBilling(
                f'Error ao tentar acessar o arquivo: {WORKBOOK_PATH_COMPANY}')
            raise exception_
        # Percorrer o sheet e pegar o nome da empresa dentro da tabela
        names_of_workbook = workbook_company.sheetnames
        company_names_list: List[Tuple[str, str]] = []

        for sheet_name in names_of_workbook:
            ws: Worksheet = workbook_company[f'{sheet_name}']
            company_name = ws['B1'].value
            if company_name:
                company_names_list.append((company_name, sheet_name))

        # Adiciona empresas do dicionário ao resultado
        for company_name in dictionary_company_names:
            company_names_list.append((company_name[1], company_name[0]))

        return company_names_list

    def test_X_Companys(self, number_companies):
        # Bloco para testar numero x de empresas
        companyList_BillingTeste = []

        for i in range(number_companies):
            companyList_BillingTeste.append(self.companyList_Billing[i])

        self.companyList_Billing = companyList_BillingTeste

        # self.companyList_Billing = [
        #     self.companyList_Billing[2], self.companyList_Billing[3]]

    async def generate_base_data_company(self):
        billing_list = self.extract_billing_list()
        self.company_list_billing = []

        for index, billing_row in enumerate(billing_list):
            self.check_companies(billing_row)
            if len(billing_list) > 0:
                progress_percent = int(30 * (index + 1) / len(billing_list))
                self.progressed.emit(10 + progress_percent)

        self.names_of_companies = self.get_name_company_in_sheet()
        self.companies_not_found.append('\nEmpresas não encontradas: ')

    def saveDictionaryExams(self):
        pathFile = resource_path('./files/dictionary_exams.json')

        with open(pathFile, 'w', encoding='utf8') as arquivo:
            json.dump(
                self.dictionary_exams,
                arquivo,
                ensure_ascii=False,
                indent=4,
            )

    def getDictionaryExams(self) -> List[tuple]:
        newDictionary = []
        pathFile = './files/dictionary_exams.json'
        with open(pathFile, 'r', encoding='utf8') as arquivo:
            newDictionary = json.load(arquivo)
        self.dictionary_exams = newDictionary
        return newDictionary

    async def create_sheet(self, company, monthText, monthNumber):
        """
        Cria uma planilha Excel para a empresa especificada.

        Args:
            company (Company): Objeto Company contendo os dados da empresa.
            monthText (str): Nome do mês.
            monthNumber (int): Número do mês.
        """
        wb: Workbook = Workbook()
        styleXlsx(wb)
        wb.remove(wb['Sheet'])
        data_company = []

        if (self.is_detail):
            data_company = setContentDetailDataFrame(company.employees)
        else:
            data_company = setContentDataFrame(company.employees)

        ws: Worksheet = wb.create_sheet(company.name[:15])

        mountStyleHeaderTable(self.yearText, monthNumber, ws, company.name)
        for item in data_company:

            if (self.is_detail):
                addDetailDataFrame(ws, item)
            else:
                # Caso eu não tenha exames encontrados
                # para esses empregado
                if (item[5] == 0):
                    item[3] = '-'
                    item[5] = '-'
                ws.append(item)
        if (company.missingExams != ''):
            showMissingExams(ws, company.missingExams)
        mountStyleContentTable(ws)
        setTotalCostRow(ws)
        resizeTable(ws)

        name_file = f'{company.name} - {monthNumber} {monthText}'
        folder_and_name = f'{self.folder}/Faturamento - {name_file}'
        wb.save(f'{folder_and_name} {self.yearText}.xlsx')

    def check_companies(self, billing_row):
        noHasCompany = True

        # Para cada empresa criar uma instancia dela com o emprego caso ja
        # exista adicionar apenas o empregado a ela.
        for company in self.companyList_Billing:
            newCompanyBilling = CompanyBilling()
            newCompanyBilling.employeesBilling = []
            employeesBillingAux = EmployeesBilling()
            if (billing_row.company_name == company.name):
                employeesBillingAux = EmployeesBilling(
                    billing_row.employee_name,
                    billing_row.employee_function,
                    billing_row.exams,
                    data=billing_row.dataExam,
                    exam_type=billing_row.exam_type)
                company.employeesBilling.append(employeesBillingAux)

                noHasCompany = False
        if (noHasCompany):
            employeesBillingAux = EmployeesBilling(
                name=billing_row.employee_name,
                function=billing_row.employee_function,
                exams=billing_row.exams,
                data=billing_row.dataExam,
                exam_type=billing_row.exam_type)
            newCompanyBilling = CompanyBilling(
                name=billing_row.company_name,
                employeesBilling=[employeesBillingAux])
            self.companyList_Billing.append(newCompanyBilling)
        if (len(self.companyList_Billing) == 0):
            newCompanyBilling = CompanyBilling()
            employeesBillingAux = EmployeesBilling(
                name=billing_row.employee_name,
                function=billing_row.employee_function,
                exams=billing_row.exams,
                data=billing_row.dataExam,
                exam_type=billing_row.exam_type)
            newCompanyBilling = CompanyBilling(
                name=billing_row.company_name,
                employeesBilling=[employeesBillingAux])
            self.companyList_Billing.append(newCompanyBilling)

    def get_company_list(self, company_Billing) -> List[Company]:
        """
        Constrói uma lista de objetos Company a partir dos dados de
        faturamento.

        Args:
            company_Billing (CompanyBilling): Objeto CompanyBilling contendo
            os dados de faturamento.

        Returns:
            List[Company]: Lista de objetos Company.
        """
        companyListAux = []
        newCompany = Company(company_Billing.name)
        newCompany.employees = []
        company_name_billing = self._normalize_name(company_Billing.name)
        hasName, company_name = self._find_company_name(company_name_billing)

        if not hasName:
            self.companies_not_found.append('\n' + company_Billing.name)
        else:
            list_of_company_exams = []
            getExams(company_name, list_of_company_exams, newCompany)

            namesAndExams = self._get_names_and_exams(company_Billing)

            for nameAndExam in namesAndExams:
                employee_company = self._create_employee(nameAndExam)
                exams_exact = nameAndExam[1].split('/')
                self._process_exams(
                    exams_exact, list_of_company_exams,
                    employee_company, newCompany)
                newCompany.employees.append(employee_company)
            companyListAux.append(newCompany)
        return companyListAux

    def _normalize_name(self, name):
        return unidecode(name.lower()) \
            .replace('ltda', '').replace('eireli', '')

    def _find_company_name(self, company_name_billing):
        for names in self.names_of_companies:
            realNameOfCompany = self._normalize_name(names[0])
            nameOfPageSheetCompany = self._normalize_name(names[1])

            if realNameOfCompany in company_name_billing or \
                    nameOfPageSheetCompany in company_name_billing:
                return True, names[1]
        return False, ''

    def _get_names_and_exams(self, company_Billing):
        namesAndExams = []
        for employees in company_Billing.employeesBilling:
            namesAndExams.append(
                (employees.name, employees.exams, employees.function,
                 employees.data, employees.exam_type))
        return namesAndExams

    def _create_employee(self, nameAndExam):
        employee_company = Employee()
        employee_company.name = nameAndExam[0]
        employee_company.function = nameAndExam[2]
        employee_company.exams = nameAndExam[1]
        employee_company.exam_type = nameAndExam[4]
        employee_company.date = setDateXlsxToString(nameAndExam[3])
        employee_company.exams_cost = []
        employee_company.cost = 0.0
        return employee_company

    def _process_exams(self, exams_exact, list_of_company,
                       employee_company, newCompany):
        """
        Processa os exames de um funcionário, atualizando os custos e os 
        exames faltantes.

        Args:
            exams_exact (list): Lista de exames.
            list_of_company (list): Lista de exames da empresa.
            employee_company (Employee): Objeto Employee do funcionário.
            newCompany (Company): Objeto Company da empresa.
        """
        for exam in exams_exact:
            examWithoutFormat = exam
            exam = self._normalize_name(exam.strip())
            exam = self._map_exam_to_dictionary(exam)
            hasExam = False
            for name_exam in list_of_company:
                if self._is_exam_in_company(exam, name_exam) and not hasExam:
                    if self._update_employee_costs(
                            name_exam, employee_company, examWithoutFormat):
                        hasExam = True

            if not hasExam:
                self._update_missing_exams(newCompany, exam)

    def _map_exam_to_dictionary(self, exam):
        for exam_significant in self.dictionary_exams:
            if unidecode(exam_significant[0].lower()) == exam:
                return exam_significant[1]
        return exam

    def _is_exam_in_company(self, exam, company):
        return unidecode(exam.lower()) in unidecode(company[0].lower())

    def _update_employee_costs(self, company, employee_company,
                               examWithoutFormat):
        if 'externo' not in examWithoutFormat.lower() and \
                'externo' not in company[0].lower():

            if self.is_detail:
                employee_company.exams_cost.append(
                    (examWithoutFormat, company[1]))
            if isinstance(company[1], (float, int)):
                employee_company.cost += company[1]
                return True
        return False

    def _update_missing_exams(self, newCompany, exam):
        if not newCompany.missingExams:
            newCompany.missingExams = exam
        elif exam not in newCompany.missingExams:
            newCompany.missingExams += ", " + exam

    # ----------------------------------------------------------------

    async def get_all_exams(self, companyList):
        locale.setlocale(locale.LC_ALL, 'pt_BR')
        monthText: str = self.monthText
        monthNumber = datetime.strptime(monthText, '%B').month

        for i, company in enumerate(companyList):
            if len(company) > 0:
                # Incrementa o progresso para cada empresa processada
                self.maxEmployees += len(company[0].employees)
                await self.create_sheet(company[0], monthText.upper(),
                                        monthNumber)
                self.progress += 1
                # Emit progress update
                progress_percent = int(30 * (i + 1) / len(companyList))
                self.progressed.emit(60 + progress_percent)

    def build_companies_list(self, companies):
        """
        Constrói a lista de empresas com os resultados do faturamento.

        Args:
            companies (list): Lista de objetos Company.

        Returns:
            list: Lista de dicionários com o nome da empresa
            e os exames faltantes.
        """
        companies_list = []
        for company in companies:
            if company:
                companies_list.append({
                    'name': company[0].name,
                    'missingExams': company[0].missingExams,
                })
        return companies_list

    async def generate_files(self):
        """
        Gera os arquivos de faturamento para todas as empresas na lista de
        faturamento.

        Returns:
            list: Lista de empresas processadas.
        """
        companyListAux = []
        for i, companyBilling in enumerate(self.companyList_Billing):
            companyListAux.append(self.get_company_list(companyBilling))
            # Emit progress update
            if len(self.companyList_Billing) > 0:
                progress_percent = int(
                    20 * (i + 1) / len(self.companyList_Billing))
                self.progressed.emit(30 + progress_percent)
        return companyListAux

    def create_output_folder(self):
        """
        Cria a pasta de saída para os arquivos de faturamento.
        """
        try:
            os.mkdir(self.folder)
        except FileExistsError:
            pass
        except Exception as e:
            raise ErrorBilling(f'Erro ao criar pasta {self.folder}: {str(e)}')

    def process_billing(self):
        """
        Processa todo o faturamento.

        Este método é responsável por coordenar todo o processo de faturamento,
        incluindo o carregamento de dados, a geração de arquivos e
        a atualização da barra de progresso.

        Emits:
            Signal: Vários sinais são emitidos para atualizar o progresso e
            indicar o início e o fim do processo.
        """
        try:
            self.reset()  # Resetando o estado antes de iniciar o processamento
            self.started.emit('Processo iniciado...')
            self.progress = 0

            # Inicializa a barra de progresso
            self.range_progress.emit(100)
            self.progressed.emit(0)

            # Carregando dados do arquivo de exames realizados
            self.load_exam_data()
            self.progressed.emit(10)

            # Gerando dados base da empresa
            asyncio.run(self.generate_base_data_company())
            self.progressed.emit(30)

            # Gerando arquivos de faturamento
            companies = asyncio.run(self.generate_files())
            self.progressed.emit(50)

            # Criando pasta que terá os arquivos de faturamentos
            self.create_output_folder()
            self.progressed.emit(60)

            # Criando os arquivos de faturamentos e populando.
            asyncio.run(self.get_all_exams(companies))
            companies_list = self.build_companies_list(companies)
            self.progressed.emit(90)

            self.finished.emit(companies_list)
            if len(self.companies_not_found) > 1:
                self.companies_not_found_signal.emit(self.companies_not_found)
            else:
                self.companies_not_found_signal.emit(
                    ['\nTodas as empresa foram geradas'])

            # Finaliza a barra de progresso
            self.progressed.emit(100)

        except ErrorBilling as error_billing:
            logging.error(f'Erro durante a geração: {str(error_billing)}')
            self.finished.emit(['Error: ', str(error_billing)])
            raise ErrorBilling('Ocorreu um error durante a geração')


# Parâmetros necessários para executar o script independentemente.
parser = ArgumentParser()

parser.add_argument("-y", "--year",
                    type=str,
                    help='Repassar o ano para a classe Billing')
parser.add_argument("-m", "--month",
                    type=str,
                    help='Repassar o mês para a classe Billing')
parser.add_argument("-f", "--folder",
                    type=str,
                    help='Repassa a pasta gerada para a classe Billing')
parser.add_argument("-d", "--detail",
                    action='store_true',
                    help='Repassa se as planilhas devem ser detalhadas')

args = parser.parse_args()

if __name__ == '__main__':
    if (len(sys.argv) > 1):
        processor = BillingDataProcessor()
        isDetail = False
        if (args.detail == 'True'):
            isDetail = True
        processor.setParamsBilling(args.year, args.month,
                                   args.folder, isDetail)

        processor.process_billing()
